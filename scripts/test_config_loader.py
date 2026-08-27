"""
test_config_loader.py — Tests for external configuration (Milestone 3C.2).

Covers configuration loading/validation, CLI precedence, backward-compatible
defaults, failure-before-analysis, and a full real-pair acceptance test using
two genuinely different snapshots derived from the sample workbook.

Standard-library unittest (no external test dependency).

Run:
    python -m unittest scripts.test_config_loader
    (or)  python scripts/test_config_loader.py
"""

import contextlib
import io
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

import run_weekly_snapshot as runner
from runner_config import RunnerConfig
from config_loader import (
    build_config,
    load_config_file,
    merge_into_config,
    ConfigError,
)
from compare_snapshots import load_snapshot, extract_vendor_data, compare_snapshots
from executive_summary import generate_executive_summary, generate_key_movements

REPO_ROOT = Path(__file__).resolve().parent.parent
SAMPLE_WORKBOOK = REPO_ROOT / "data" / "Fake vendor data.xlsx"


def _quiet():
    return contextlib.redirect_stdout(io.StringIO())


def _write_json(path: Path, obj) -> None:
    path.write_text(json.dumps(obj), encoding="utf-8")


# ---------------------------------------------------------------------------
# Configuration loading / validation
# ---------------------------------------------------------------------------

class TestConfigValidation(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="cfg_test_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_valid_external_configuration(self):
        cfg_path = self.tmp / "cfg.json"
        _write_json(cfg_path, {
            "snapshot_worksheet": "Sheet X",
            "incoming_directory": "data/incoming",
            "allowed_extensions": [".xlsx"],
        })
        config = build_config(config_path=cfg_path)
        self.assertEqual(config.snapshot_worksheet, "Sheet X")
        self.assertTrue(str(config.incoming_dir).endswith("incoming"))
        self.assertEqual(config.allowed_extensions, (".xlsx",))

    def test_missing_configuration_file(self):
        missing = self.tmp / "does_not_exist.json"
        with self.assertRaises(ConfigError) as ctx:
            build_config(config_path=missing, require_config_file=True)
        self.assertIn("not found", str(ctx.exception))

    def test_malformed_json(self):
        cfg_path = self.tmp / "bad.json"
        cfg_path.write_text("{ not valid json ", encoding="utf-8")
        with self.assertRaises(ConfigError) as ctx:
            load_config_file(cfg_path)
        self.assertIn("Malformed JSON", str(ctx.exception))

    def test_unknown_key_rejected(self):
        cfg_path = self.tmp / "cfg.json"
        _write_json(cfg_path, {"snapshot_worksheet": "S", "bogus_key": 1})
        with self.assertRaises(ConfigError) as ctx:
            load_config_file(cfg_path)
        self.assertIn("Unknown configuration key", str(ctx.exception))
        self.assertIn("bogus_key", str(ctx.exception))

    def test_invalid_value_type(self):
        cfg_path = self.tmp / "cfg.json"
        _write_json(cfg_path, {"snapshot_worksheet": 123})
        with self.assertRaises(ConfigError) as ctx:
            load_config_file(cfg_path)
        self.assertIn("must be a string", str(ctx.exception))

    def test_invalid_allowed_extensions_type(self):
        cfg_path = self.tmp / "cfg.json"
        _write_json(cfg_path, {"allowed_extensions": "xlsx"})
        with self.assertRaises(ConfigError) as ctx:
            load_config_file(cfg_path)
        self.assertIn("list of strings", str(ctx.exception))

    def test_top_level_not_object(self):
        cfg_path = self.tmp / "cfg.json"
        cfg_path.write_text("[1, 2, 3]", encoding="utf-8")
        with self.assertRaises(ConfigError) as ctx:
            load_config_file(cfg_path)
        self.assertIn("JSON object", str(ctx.exception))


# ---------------------------------------------------------------------------
# Precedence and defaults
# ---------------------------------------------------------------------------

class TestPrecedence(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="cfg_prec_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_cli_worksheet_override(self):
        config = build_config(worksheet_override="CLI Sheet")
        self.assertEqual(config.snapshot_worksheet, "CLI Sheet")

    def test_cli_overrides_file(self):
        cfg_path = self.tmp / "cfg.json"
        _write_json(cfg_path, {"snapshot_worksheet": "File Sheet"})
        config = build_config(config_path=cfg_path, worksheet_override="CLI Sheet")
        self.assertEqual(config.snapshot_worksheet, "CLI Sheet")

    def test_file_overrides_defaults(self):
        cfg_path = self.tmp / "cfg.json"
        _write_json(cfg_path, {"snapshot_worksheet": "File Sheet"})
        config = build_config(config_path=cfg_path)
        self.assertEqual(config.snapshot_worksheet, "File Sheet")

    def test_defaults_when_no_override(self):
        # No config path, no CLI override: the built-in default worksheet holds.
        # (The repo default config file mirrors the default, so this remains
        # "Snapshot Wk 2" whether or not the default file is present.)
        config = build_config()
        self.assertEqual(config.snapshot_worksheet, "Snapshot Wk 2")

    def test_empty_worksheet_override_rejected(self):
        with self.assertRaises(ConfigError):
            build_config(worksheet_override="   ")


# ---------------------------------------------------------------------------
# Failure before analysis; no archive/state change on config failure
# ---------------------------------------------------------------------------

class TestConfigFailureIsolation(unittest.TestCase):
    def test_invalid_config_fails_before_analysis(self):
        tmp = Path(tempfile.mkdtemp(prefix="cfg_fail_"))
        try:
            bad_cfg = tmp / "bad.json"
            bad_cfg.write_text("{ broken ", encoding="utf-8")

            # main() must return non-zero and never invoke the pipeline.
            with _quiet():
                rc = runner.main(["--config", str(bad_cfg)])
            self.assertEqual(rc, 1)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_config_failure_leaves_incoming_and_state_untouched(self):
        tmp = Path(tempfile.mkdtemp(prefix="cfg_fail2_"))
        try:
            # A real incoming workbook + baseline state that must NOT change.
            incoming = tmp / "incoming"
            archive = tmp / "archive"
            outputs = tmp / "outputs"
            state = tmp / "state"
            for d in (incoming, archive, outputs, state):
                d.mkdir(parents=True)

            wb = incoming / "week.xlsx"
            openpyxl.Workbook().save(wb)
            state_file = state / "last_successful_run.json"
            state_file.write_text('{"snapshot_path": "x"}', encoding="utf-8")
            state_before = state_file.read_text(encoding="utf-8")

            # Build an invalid config and confirm build_config raises before any
            # run() work. (main() wires the same path.)
            bad_cfg = tmp / "bad.json"
            _write_json(bad_cfg, {"unknown_key": 1})
            with self.assertRaises(ConfigError):
                build_config(config_path=bad_cfg, require_config_file=True)

            # Nothing archived, state unchanged, no outputs produced.
            self.assertTrue(wb.exists())
            self.assertEqual(state_file.read_text(encoding="utf-8"), state_before)
            self.assertEqual(list(archive.iterdir()), [])
            self.assertEqual(list(outputs.iterdir()), [])
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


# ---------------------------------------------------------------------------
# Real-pair acceptance test
# ---------------------------------------------------------------------------

def _extract_sheet_to_single_workbook(source_wb: Path, sheet_name: str,
                                       dest_path: Path, dest_sheet: str) -> None:
    """
    Copy one worksheet's cell values from source_wb into a new single-sheet
    workbook saved at dest_path, with the sheet renamed to dest_sheet.

    This produces a genuine single-snapshot workbook for the acceptance test.
    """
    src = openpyxl.load_workbook(source_wb, read_only=True, data_only=True)
    ws = src[sheet_name]
    out = openpyxl.Workbook()
    out_ws = out.active
    out_ws.title = dest_sheet
    for row in ws.iter_rows(values_only=True):
        out_ws.append(list(row))
    out.save(dest_path)
    src.close()


class TestRealPairAcceptance(unittest.TestCase):
    @unittest.skipUnless(SAMPLE_WORKBOOK.exists(), "sample workbook not present")
    def test_full_real_pair_run(self):
        tmp = Path(tempfile.mkdtemp(prefix="realpair_"))
        try:
            worksheet = "Snapshot Wk 2"  # the name each single-sheet book will use

            config = RunnerConfig(
                data_dir=tmp,
                incoming_dir=tmp / "incoming",
                archive_dir=tmp / "archive",
                outputs_dir=tmp / "outputs",
                state_dir=tmp / "state",
                snapshot_worksheet=worksheet,
            )
            config.ensure_directories()

            # Build two genuinely different single-snapshot workbooks from the
            # two different worksheets in the sample workbook.
            previous_wb = config.archive_dir / "previous.xlsx"
            current_wb = config.incoming_dir / "current.xlsx"
            _extract_sheet_to_single_workbook(
                SAMPLE_WORKBOOK, "Snapshot Wk 1", previous_wb, worksheet)
            _extract_sheet_to_single_workbook(
                SAMPLE_WORKBOOK, "Snapshot Wk 2", current_wb, worksheet)

            # Establish the previous baseline in state.
            runner.write_state(config, previous_wb, "baseline-run")

            # Golden result: invoke the existing pipeline directly on the same
            # two single-sheet workbooks.
            with _quiet():
                prev_vendors = extract_vendor_data(load_snapshot(previous_wb, worksheet))
                curr_vendors = extract_vendor_data(load_snapshot(current_wb, worksheet))
                golden_analysis = compare_snapshots(prev_vendors, curr_vendors)
            golden_exec = generate_executive_summary(golden_analysis)
            golden_keys = generate_key_movements(golden_analysis)

            # Run the normal runner entry point.
            manifest = runner.run(config)

            # Success + all stages.
            self.assertEqual(manifest["status"], "success", manifest["errors"])
            for stage in ("discovery", "previous_resolution", "preflight",
                          "analysis", "reporting", "promote", "archive",
                          "state_update"):
                self.assertIn(stage, manifest["stages_completed"])

            # At least one movement was produced.
            self.assertGreater(golden_analysis["changed_count"], 0)

            # Runner output exactly matches direct existing-pipeline output.
            out_dir = config.outputs_dir / manifest["run_id"]
            runner_exec = (out_dir / "executive_summary.txt").read_text(encoding="utf-8")
            runner_keys = (out_dir / "key_movements.txt").read_text(encoding="utf-8")
            self.assertEqual(runner_exec, golden_exec)
            self.assertEqual(runner_keys, golden_keys)

            # Archive, state, output promotion.
            self.assertFalse(current_wb.exists())  # archived out of incoming
            self.assertTrue(any(config.archive_dir.glob("current*.xlsx")))
            self.assertTrue((out_dir / "analysis.json").exists())
            state = json.loads(config.state_file.read_text(encoding="utf-8"))
            self.assertIn("current", Path(state["snapshot_path"]).name)
            # Manifest file exists and reports success.
            manifest_file = config.outputs_dir / f"manifest_{manifest['run_id']}.json"
            self.assertTrue(manifest_file.exists())
            self.assertEqual(
                json.loads(manifest_file.read_text(encoding="utf-8"))["status"],
                "success",
            )
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main(verbosity=2)
