"""
test_weekly_update.py — Tests for Weekly Update Assembly (Milestone 3D.1).

Standard-library unittest (no external test dependency).

Run:
    python -m unittest reporting.test_weekly_update
    (or)  python reporting/test_weekly_update.py
"""

import contextlib
import io
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from reporting.weekly_update import (
    assemble_weekly_update,
    render_weekly_update,
    WeeklyUpdateError,
    WEEKLY_UPDATE_FILENAME,
)

SAMPLE_WORKBOOK = REPO_ROOT / "data" / "Fake vendor data.xlsx"

EXEC_CONTENT = "EXEC SUMMARY LINE 1\nEXEC SUMMARY LINE 2"
MOVES_CONTENT = "KEY MOVEMENTS LINE 1\nKEY MOVEMENTS LINE 2"


def _quiet():
    return contextlib.redirect_stdout(io.StringIO())


def _seed_artefacts(directory: Path, exec_content=EXEC_CONTENT,
                    moves_content=MOVES_CONTENT) -> None:
    (directory / "executive_summary.txt").write_text(exec_content, encoding="utf-8")
    (directory / "key_movements.txt").write_text(moves_content, encoding="utf-8")


class TestAssembleUnit(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="wu_unit_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_successful_assembly(self):
        _seed_artefacts(self.tmp)
        out = assemble_weekly_update(
            output_dir=self.tmp,
            run_id="RUN-1",
            previous_snapshot="prev.xlsx",
            current_snapshot="curr.xlsx",
            generated_timestamp="2026-01-01T00:00:00+00:00",
        )
        self.assertEqual(out.name, WEEKLY_UPDATE_FILENAME)
        self.assertTrue(out.exists())
        text = out.read_text(encoding="utf-8")
        # Fixed sections present and ordered.
        for section in ("# Weekly IT Simplification Update", "## Run Information",
                        "## Executive Summary", "## Key Movements"):
            self.assertIn(section, text)
        positions = [text.index(s) for s in (
            "# Weekly IT Simplification Update", "## Run Information",
            "## Executive Summary", "## Key Movements")]
        self.assertEqual(positions, sorted(positions))
        # Metadata present.
        self.assertIn("Run ID: RUN-1", text)
        self.assertIn("Previous Snapshot: prev.xlsx", text)
        self.assertIn("Current Snapshot: curr.xlsx", text)
        self.assertIn("Generated: 2026-01-01T00:00:00+00:00", text)
        # Horizontal-rule separators between sections.
        self.assertEqual(text.count("\n---\n"), 2)
        # Upstream content preserved verbatim.
        self.assertIn(EXEC_CONTENT, text)
        self.assertIn(MOVES_CONTENT, text)

    def test_missing_summary_fails_fast(self):
        # Only movements present.
        (self.tmp / "key_movements.txt").write_text(MOVES_CONTENT, encoding="utf-8")
        with self.assertRaises(WeeklyUpdateError) as ctx:
            assemble_weekly_update(
                output_dir=self.tmp, run_id="R", previous_snapshot="p",
                current_snapshot="c", generated_timestamp="t",
            )
        self.assertIn("executive summary", str(ctx.exception))
        # No weekly_update.md produced.
        self.assertFalse((self.tmp / WEEKLY_UPDATE_FILENAME).exists())

    def test_missing_movements_fails_fast(self):
        # Only summary present.
        (self.tmp / "executive_summary.txt").write_text(EXEC_CONTENT, encoding="utf-8")
        with self.assertRaises(WeeklyUpdateError) as ctx:
            assemble_weekly_update(
                output_dir=self.tmp, run_id="R", previous_snapshot="p",
                current_snapshot="c", generated_timestamp="t",
            )
        self.assertIn("key movements", str(ctx.exception))
        self.assertFalse((self.tmp / WEEKLY_UPDATE_FILENAME).exists())

    def test_deterministic_output(self):
        _seed_artefacts(self.tmp)
        kwargs = dict(
            output_dir=self.tmp, run_id="RUN-9", previous_snapshot="p.xlsx",
            current_snapshot="c.xlsx", generated_timestamp="2026-05-05T12:00:00+00:00",
        )
        first = assemble_weekly_update(**kwargs).read_text(encoding="utf-8")
        second = assemble_weekly_update(**kwargs).read_text(encoding="utf-8")
        self.assertEqual(first, second)
        # render function is also pure/deterministic for identical inputs.
        r1 = render_weekly_update(EXEC_CONTENT, MOVES_CONTENT, "RUN-9",
                                  "p.xlsx", "c.xlsx", "2026-05-05T12:00:00+00:00")
        r2 = render_weekly_update(EXEC_CONTENT, MOVES_CONTENT, "RUN-9",
                                  "p.xlsx", "c.xlsx", "2026-05-05T12:00:00+00:00")
        self.assertEqual(r1, r2)


def _extract_sheet_to_single_workbook(source_wb, sheet_name, dest_path, dest_sheet):
    src = openpyxl.load_workbook(source_wb, read_only=True, data_only=True)
    ws = src[sheet_name]
    out = openpyxl.Workbook()
    out_ws = out.active
    out_ws.title = dest_sheet
    for row in ws.iter_rows(values_only=True):
        out_ws.append(list(row))
    out.save(dest_path)
    src.close()


class TestAcceptanceRealRun(unittest.TestCase):
    @unittest.skipUnless(SAMPLE_WORKBOOK.exists(), "sample workbook not present")
    def test_weekly_update_produced_by_real_run(self):
        # Import the runner here so this module can be run standalone.
        import run_weekly_snapshot as runner
        from runner_config import RunnerConfig

        tmp = Path(tempfile.mkdtemp(prefix="wu_accept_"))
        try:
            worksheet = "Snapshot Wk 2"
            config = RunnerConfig(
                data_dir=tmp,
                incoming_dir=tmp / "incoming",
                archive_dir=tmp / "archive",
                outputs_dir=tmp / "outputs",
                state_dir=tmp / "state",
                snapshot_worksheet=worksheet,
            )
            config.ensure_directories()

            previous_wb = config.archive_dir / "previous.xlsx"
            current_wb = config.incoming_dir / "current.xlsx"
            _extract_sheet_to_single_workbook(
                SAMPLE_WORKBOOK, "Snapshot Wk 1", previous_wb, worksheet)
            _extract_sheet_to_single_workbook(
                SAMPLE_WORKBOOK, "Snapshot Wk 2", current_wb, worksheet)
            runner.write_state(config, previous_wb, "baseline-run")

            manifest = runner.run(config)

            self.assertEqual(manifest["status"], "success", manifest["errors"])
            self.assertIn("weekly_update", manifest["stages_completed"])

            out_dir = config.outputs_dir / manifest["run_id"]
            weekly = out_dir / WEEKLY_UPDATE_FILENAME
            self.assertTrue(weekly.exists())

            text = weekly.read_text(encoding="utf-8")
            # Metadata from the real run.
            self.assertIn(f"Run ID: {manifest['run_id']}", text)
            self.assertIn("Previous Snapshot:", text)
            self.assertIn("Current Snapshot:", text)
            self.assertIn("Generated:", text)
            # Verbatim inclusion of the two real artefacts.
            exec_txt = (out_dir / "executive_summary.txt").read_text(encoding="utf-8")
            moves_txt = (out_dir / "key_movements.txt").read_text(encoding="utf-8")
            self.assertIn(exec_txt, text)
            self.assertIn(moves_txt, text)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main(verbosity=2)
