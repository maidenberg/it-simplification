"""
test_promotion_package.py — Tests for Promotion Package assembly (Milestone 3D.5).

Standard-library unittest (no external test dependency).

Run:
    python -m unittest tests.test_promotion_package
"""

import contextlib
import io
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from src.reporting.promotion_package import (
    generate_promotion_package,
    PromotionPackageError,
    REQUIRED_ARTEFACTS,
    PROMOTION_PACKAGE_DIRNAME,
    PROMOTION_MANIFEST_FILENAME,
)

SAMPLE_WORKBOOK = REPO_ROOT / "data" / "Fake vendor data.xlsx"


def _quiet():
    return contextlib.redirect_stdout(io.StringIO())


class PromotionPackageTestBase(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="pp_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _seed_all(self):
        for i, name in enumerate(REQUIRED_ARTEFACTS):
            (self.tmp / name).write_text(f"content of {name}\nline 2 {i}", encoding="utf-8")

    def _generate(self, run_id="RUN-1", generated_at="2026-01-01T00:00:00+00:00"):
        return generate_promotion_package(
            source_dir=self.tmp, run_id=run_id, generated_at=generated_at
        )


class TestSuccessPath(PromotionPackageTestBase):
    def test_success_creates_package_and_copies_files(self):
        self._seed_all()
        pkg = self._generate()
        self.assertTrue(pkg.exists())
        self.assertEqual(pkg.name, PROMOTION_PACKAGE_DIRNAME)
        # All artefacts present in the package.
        for name in REQUIRED_ARTEFACTS:
            self.assertTrue((pkg / name).exists())
        # Manifest present.
        self.assertTrue((pkg / PROMOTION_MANIFEST_FILENAME).exists())

    def test_files_copied_unchanged(self):
        self._seed_all()
        pkg = self._generate()
        for name in REQUIRED_ARTEFACTS:
            src = (self.tmp / name).read_text(encoding="utf-8")
            dst = (pkg / name).read_text(encoding="utf-8")
            self.assertEqual(src, dst)


class TestManifest(PromotionPackageTestBase):
    def test_manifest_generation(self):
        self._seed_all()
        pkg = self._generate(run_id="RUN-XYZ", generated_at="2026-05-05T10:00:00+00:00")
        manifest = json.loads((pkg / PROMOTION_MANIFEST_FILENAME).read_text(encoding="utf-8"))
        self.assertEqual(manifest["run_id"], "RUN-XYZ")
        self.assertEqual(manifest["generated_at"], "2026-05-05T10:00:00+00:00")
        self.assertEqual(manifest["files"], list(REQUIRED_ARTEFACTS))
        # Metadata only — no analytics keys.
        self.assertEqual(set(manifest.keys()), {"run_id", "generated_at", "files"})

    def test_deterministic_output(self):
        self._seed_all()
        pkg1 = self._generate()
        m1 = (pkg1 / PROMOTION_MANIFEST_FILENAME).read_text(encoding="utf-8")
        pkg2 = self._generate()
        m2 = (pkg2 / PROMOTION_MANIFEST_FILENAME).read_text(encoding="utf-8")
        self.assertEqual(m1, m2)


class TestMissingAndEmpty(PromotionPackageTestBase):
    def test_missing_file(self):
        self._seed_all()
        (self.tmp / "leadership_insights.txt").unlink()
        with self.assertRaises(PromotionPackageError) as ctx:
            self._generate()
        self.assertIn("not found", str(ctx.exception))
        self.assertIn("leadership_insights.txt", str(ctx.exception))
        # Package not created on failure.
        self.assertFalse((self.tmp / PROMOTION_PACKAGE_DIRNAME).exists())

    def test_empty_file(self):
        self._seed_all()
        (self.tmp / "risks_watchouts.txt").write_text("", encoding="utf-8")
        with self.assertRaises(PromotionPackageError) as ctx:
            self._generate()
        self.assertIn("empty", str(ctx.exception))
        self.assertIn("risks_watchouts.txt", str(ctx.exception))


class TestExistingPackageDirectory(PromotionPackageTestBase):
    def test_existing_package_replaced_without_stale_content(self):
        self._seed_all()
        pkg = self._generate()
        # Drop a stale file into the existing package directory.
        stale = pkg / "stale_file.txt"
        stale.write_text("stale", encoding="utf-8")
        # Regenerate; stale content must be gone, required files present.
        pkg2 = self._generate()
        self.assertFalse((pkg2 / "stale_file.txt").exists())
        for name in REQUIRED_ARTEFACTS:
            self.assertTrue((pkg2 / name).exists())


def _extract_sheet(src, sheet, dest, dest_sheet):
    s = openpyxl.load_workbook(src, read_only=True, data_only=True)
    w = s[sheet]
    o = openpyxl.Workbook()
    ow = o.active
    ow.title = dest_sheet
    for row in w.iter_rows(values_only=True):
        ow.append(list(row))
    o.save(dest)
    s.close()


@unittest.skipUnless(SAMPLE_WORKBOOK.exists(), "sample workbook not present")
class TestRunnerIntegration(unittest.TestCase):
    def test_promotion_package_produced_and_ordered(self):
        import run_weekly_snapshot as runner
        from runner_config import RunnerConfig

        tmp = Path(tempfile.mkdtemp(prefix="pp_run_"))
        try:
            ws = "Snapshot Wk 2"
            config = RunnerConfig(
                data_dir=tmp, incoming_dir=tmp / "incoming",
                archive_dir=tmp / "archive", outputs_dir=tmp / "outputs",
                state_dir=tmp / "state", snapshot_worksheet=ws,
            )
            config.ensure_directories()
            prev = config.archive_dir / "prev.xlsx"
            curr = config.incoming_dir / "curr.xlsx"
            _extract_sheet(SAMPLE_WORKBOOK, "Snapshot Wk 1", prev, ws)
            _extract_sheet(SAMPLE_WORKBOOK, "Snapshot Wk 2", curr, ws)
            runner.write_state(config, prev, "baseline")

            with _quiet():
                manifest = runner.run(config)

            self.assertEqual(manifest["status"], "success", manifest["errors"])
            stages = manifest["stages_completed"]
            # Order: reporting_package -> promotion_package -> promote.
            self.assertLess(stages.index("reporting_package"),
                            stages.index("promotion_package"))
            self.assertLess(stages.index("promotion_package"),
                            stages.index("promote"))

            out_dir = config.outputs_dir / manifest["run_id"]
            pkg = out_dir / PROMOTION_PACKAGE_DIRNAME
            self.assertTrue(pkg.is_dir())
            for name in REQUIRED_ARTEFACTS:
                self.assertTrue((pkg / name).exists())
            pkg_manifest = json.loads(
                (pkg / PROMOTION_MANIFEST_FILENAME).read_text(encoding="utf-8"))
            self.assertEqual(pkg_manifest["run_id"], manifest["run_id"])
            self.assertEqual(pkg_manifest["files"], list(REQUIRED_ARTEFACTS))
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main(verbosity=2)
