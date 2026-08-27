"""
test_reporting_package_runner.py — Runner integration tests for Milestone 3D.4.

Verifies the reporting_package stage within a full runner run:
- produced by a successful run
- runs after risks_watchouts and before promote
- package failure prevents promotion (and leaves state/input untouched)
- reporting_package.txt included in the manifest generated_artefacts
- existing generated artefacts remain unchanged

Standard-library unittest.

Run:
    python -m unittest tests.test_reporting_package_runner
"""

import contextlib
import io
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "scripts"))

import run_weekly_snapshot as runner
from runner_config import RunnerConfig

SAMPLE_WORKBOOK = REPO_ROOT / "data" / "Fake vendor data.xlsx"


def _quiet():
    return contextlib.redirect_stdout(io.StringIO())


def _extract_sheet(src, sheet, dest, dest_sheet):
    s = openpyxl.load_workbook(src, read_only=True, data_only=True)
    w = s[sheet]
    o = openpyxl.Workbook()
    ow = o.active
    ow.title = dest_sheet
    for row in w.iter_rows(values_only=True):
        ow.append(list(row))
    o.save(dest)
    s.close()


@unittest.skipUnless(SAMPLE_WORKBOOK.exists(), "sample workbook not present")
class TestReportingPackageRunner(unittest.TestCase):
    def _make_config(self, tmp):
        ws = "Snapshot Wk 2"
        config = RunnerConfig(
            data_dir=tmp, incoming_dir=tmp / "incoming",
            archive_dir=tmp / "archive", outputs_dir=tmp / "outputs",
            state_dir=tmp / "state", snapshot_worksheet=ws,
        )
        config.ensure_directories()
        prev = config.archive_dir / "prev.xlsx"
        curr = config.incoming_dir / "curr.xlsx"
        _extract_sheet(SAMPLE_WORKBOOK, "Snapshot Wk 1", prev, ws)
        _extract_sheet(SAMPLE_WORKBOOK, "Snapshot Wk 2", curr, ws)
        runner.write_state(config, prev, "baseline")
        return config

    def test_reporting_package_produced_and_ordered(self):
        tmp = Path(tempfile.mkdtemp(prefix="rp_run_"))
        try:
            config = self._make_config(tmp)
            with _quiet():
                manifest = runner.run(config)

            self.assertEqual(manifest["status"], "success", manifest["errors"])
            stages = manifest["stages_completed"]
            # Order: risks_watchouts -> reporting_package -> promote.
            self.assertLess(stages.index("risks_watchouts"),
                            stages.index("reporting_package"))
            self.assertLess(stages.index("reporting_package"),
                            stages.index("promote"))

            out_dir = config.outputs_dir / manifest["run_id"]
            package = out_dir / "reporting_package.txt"
            self.assertTrue(package.exists())

            text = package.read_text(encoding="utf-8")
            for heading in ("IT SIMPLIFICATION WEEKLY REPORT", "EXECUTIVE SUMMARY",
                            "KEY MOVEMENTS", "WEEKLY UPDATE", "LEADERSHIP INSIGHTS",
                            "RISKS & WATCHOUTS", "END OF REPORT"):
                self.assertIn(heading, text)

            # Manifest includes reporting_package.txt in generated_artefacts.
            self.assertIn("reporting_package.txt", manifest["generated_artefacts"])
            manifest_file = config.outputs_dir / f"manifest_{manifest['run_id']}.json"
            recorded = json.loads(manifest_file.read_text(encoding="utf-8"))
            self.assertIn("reporting_package.txt", recorded["generated_artefacts"])

            # Existing artefacts remain present and unchanged in structure.
            for name in ("executive_summary.txt", "key_movements.txt",
                         "weekly_update.md", "leadership_insights.txt",
                         "risks_watchouts.txt", "analysis.json"):
                self.assertTrue((out_dir / name).exists())
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_package_failure_prevents_promotion(self):
        tmp = Path(tempfile.mkdtemp(prefix="rp_fail_"))
        try:
            config = self._make_config(tmp)
            state_before = config.state_file.read_text(encoding="utf-8")
            incoming_before = list(config.incoming_dir.iterdir())

            original = runner.generate_reporting_package

            def boom(**kwargs):
                raise RuntimeError("simulated package failure")

            runner.generate_reporting_package = boom
            try:
                with _quiet():
                    manifest = runner.run(config)
            finally:
                runner.generate_reporting_package = original

            self.assertEqual(manifest["status"], "failed")
            # Not promoted.
            self.assertNotIn("promote", manifest["stages_completed"])
            self.assertFalse((config.outputs_dir / manifest["run_id"]).exists())
            # Input not archived; state unchanged.
            self.assertEqual(list(config.incoming_dir.iterdir()), incoming_before)
            self.assertEqual(config.state_file.read_text(encoding="utf-8"), state_before)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main(verbosity=2)
