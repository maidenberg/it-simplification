"""
test_risks_watchouts.py — Tests for Risks & Watchouts (Milestone 3D.3).

Standard-library unittest (no external test dependency).

Run:
    python -m unittest tests.test_risks_watchouts
    (or)  python tests/test_risks_watchouts.py
"""

import shutil
import sys
import tempfile
import unittest
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from src.reporting.risks_watchouts import (
    generate_risks_watchouts,
    RisksWatchoutsError,
    NO_RISKS,
    NO_WATCHOUTS,
    NO_OBSERVATIONS,
)

# Leadership insights containing pre-labelled entries (preferred approach).
LEADERSHIP_LABELLED = "\n".join([
    "Leadership Insights",
    "",
    "RISK: Vendor X contract lapses next week.",
    "WATCHOUT: Concentration in top vendor.",
    "OBSERVATION: 142 contracts under review.",
])


LEADERSHIP_ARTEFACT = "\n".join([
    "Leadership Insights",
    "",
    "1. Portfolio review covered 142 contracts.",
    "",
    "2. Portfolio net movement for the reporting period was $24,750.09.",
    "",
    "3. Largest portfolio movement was Example Contract 028 (+$16,764.09).",
    "",
    "4. Second largest portfolio movement was Example Contract 034 (+$7,489.00).",
    "",
    "5. Top two ranked movements were Example Contract 028 and Example Contract 034.",
])

# Only positive movers -> watchouts populated, no risks.
MOVES_POSITIVE_ONLY = "\n".join([
    "KEY MOVEMENTS",
    "-------------",
    "",
    "1. Example Contract 028 (+$16,764.09)",
    "2. Example Contract 034 (+$7,489.00)",
    "3. Example Contract 023 (+$497.00)",
])

# Includes a negative mover -> risks populated.
MOVES_WITH_NEGATIVE = "\n".join([
    "KEY MOVEMENTS",
    "-------------",
    "",
    "1. Example Contract 028 (+$16,764.09)",
    "2. Example Contract 099 (-$5,560.00)",
])

# No ranked movers -> neither risks nor watchouts.
MOVES_EMPTY = "\n".join([
    "KEY MOVEMENTS",
    "-------------",
    "",
    "1.",
    "2.",
    "3.",
])


class RisksWatchoutsTestBase(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="rw_"))
        self.leadership_path = self.tmp / "leadership_insights.txt"
        self.moves_path = self.tmp / "key_movements.txt"
        self.out_path = self.tmp / "risks_watchouts.txt"

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _seed(self, leadership=LEADERSHIP_ARTEFACT, moves=MOVES_POSITIVE_ONLY):
        self.leadership_path.write_text(leadership, encoding="utf-8")
        self.moves_path.write_text(moves, encoding="utf-8")

    def _generate(self):
        return generate_risks_watchouts(
            self.leadership_path, self.moves_path, self.out_path
        )


class TestStructure(RisksWatchoutsTestBase):
    def test_file_generation_and_sections(self):
        self._seed()
        out = self._generate()
        self.assertTrue(out.exists())
        text = out.read_text(encoding="utf-8")
        for section in ("IT SIMPLIFICATION RISKS & WATCHOUTS", "RISKS",
                        "WATCHOUTS", "DATA OBSERVATIONS", "END OF REPORT"):
            self.assertIn(section, text)
        positions = [text.index(s) for s in (
            "RISKS", "WATCHOUTS", "DATA OBSERVATIONS", "END OF REPORT")]
        self.assertEqual(positions, sorted(positions))


class TestPopulated(RisksWatchoutsTestBase):
    def test_populated_watchouts(self):
        self._seed(moves=MOVES_POSITIVE_ONLY)
        text = self._generate().read_text(encoding="utf-8")
        self.assertIn("Significant movement: Example Contract 028 (+$16,764.09)", text)
        self.assertIn("Significant movement: Example Contract 034 (+$7,489.00)", text)

    def test_populated_risks(self):
        self._seed(moves=MOVES_WITH_NEGATIVE)
        text = self._generate().read_text(encoding="utf-8")
        self.assertIn("Negative movement: Example Contract 099 (-$5,560.00)", text)
        # The positive mover still appears as a watchout.
        self.assertIn("Significant movement: Example Contract 028 (+$16,764.09)", text)

    def test_data_observations_sourced_from_leadership(self):
        self._seed()
        text = self._generate().read_text(encoding="utf-8")
        self.assertIn("1. Portfolio review covered 142 contracts.", text)
        self.assertIn(
            "2. Portfolio net movement for the reporting period was $24,750.09.", text
        )


class TestEmptyPlaceholders(RisksWatchoutsTestBase):
    def test_no_risks_placeholder(self):
        # Positive-only movers => no risks.
        self._seed(moves=MOVES_POSITIVE_ONLY)
        text = self._generate().read_text(encoding="utf-8")
        self.assertIn(NO_RISKS, text)

    def test_no_watchouts_placeholder(self):
        # Empty movers => no watchouts (and no risks).
        self._seed(moves=MOVES_EMPTY)
        text = self._generate().read_text(encoding="utf-8")
        self.assertIn(NO_WATCHOUTS, text)
        self.assertIn(NO_RISKS, text)

    def test_no_observations_placeholder(self):
        # Leadership insights without insight 1/2 lines => no observations.
        self._seed(leadership="Leadership Insights\n\n(no numbered facts)",
                   moves=MOVES_POSITIVE_ONLY)
        text = self._generate().read_text(encoding="utf-8")
        self.assertIn(NO_OBSERVATIONS, text)


class TestPreferredLabelledApproach(RisksWatchoutsTestBase):
    def test_prelabelled_entries_take_precedence(self):
        # When leadership_insights carries labelled entries, they are used
        # verbatim in preference to deriving from key_movements.
        self._seed(leadership=LEADERSHIP_LABELLED, moves=MOVES_WITH_NEGATIVE)
        text = self._generate().read_text(encoding="utf-8")
        self.assertIn("Vendor X contract lapses next week.", text)
        self.assertIn("Concentration in top vendor.", text)
        self.assertIn("142 contracts under review.", text)
        # The key_movements fallback derivation is NOT used when labels exist.
        self.assertNotIn("Negative movement:", text)
        self.assertNotIn("Significant movement:", text)


class TestDeterminism(RisksWatchoutsTestBase):
    def test_deterministic_output(self):
        self._seed(moves=MOVES_WITH_NEGATIVE)
        first = self._generate().read_text(encoding="utf-8")
        second = self._generate().read_text(encoding="utf-8")
        self.assertEqual(first, second)


class TestFailureHandling(RisksWatchoutsTestBase):
    def test_missing_leadership_raises(self):
        self.moves_path.write_text(MOVES_POSITIVE_ONLY, encoding="utf-8")
        with self.assertRaises(RisksWatchoutsError) as ctx:
            self._generate()
        self.assertIn("leadership insights", str(ctx.exception))
        self.assertFalse(self.out_path.exists())

    def test_missing_key_movements_raises(self):
        self.leadership_path.write_text(LEADERSHIP_ARTEFACT, encoding="utf-8")
        with self.assertRaises(RisksWatchoutsError) as ctx:
            self._generate()
        self.assertIn("key movements", str(ctx.exception))
        self.assertFalse(self.out_path.exists())


class TestRunnerIntegration(unittest.TestCase):
    SAMPLE_WORKBOOK = REPO_ROOT / "data" / "Fake vendor data.xlsx"

    @unittest.skipUnless((REPO_ROOT / "data" / "Fake vendor data.xlsx").exists(),
                         "sample workbook not present")
    def test_runner_produces_risks_watchouts(self):
        import openpyxl
        sys.path.insert(0, str(REPO_ROOT / "scripts"))
        import run_weekly_snapshot as runner
        from runner_config import RunnerConfig

        def _extract(src, sheet, dest, ds):
            s = openpyxl.load_workbook(src, read_only=True, data_only=True)
            w = s[sheet]
            o = openpyxl.Workbook()
            ow = o.active
            ow.title = ds
            for row in w.iter_rows(values_only=True):
                ow.append(list(row))
            o.save(dest)
            s.close()

        tmp = Path(tempfile.mkdtemp(prefix="rw_run_"))
        try:
            ws = "Snapshot Wk 2"
            config = RunnerConfig(
                data_dir=tmp, incoming_dir=tmp / "incoming",
                archive_dir=tmp / "archive", outputs_dir=tmp / "outputs",
                state_dir=tmp / "state", snapshot_worksheet=ws,
            )
            config.ensure_directories()
            prev = config.archive_dir / "prev.xlsx"
            curr = config.incoming_dir / "curr.xlsx"
            _extract(self.SAMPLE_WORKBOOK, "Snapshot Wk 1", prev, ws)
            _extract(self.SAMPLE_WORKBOOK, "Snapshot Wk 2", curr, ws)
            runner.write_state(config, prev, "baseline")

            manifest = runner.run(config)

            self.assertEqual(manifest["status"], "success", manifest["errors"])
            # Ordering: risks_watchouts runs after leadership_insights, before promote.
            stages = manifest["stages_completed"]
            self.assertLess(stages.index("leadership_insights"),
                            stages.index("risks_watchouts"))
            self.assertLess(stages.index("risks_watchouts"),
                            stages.index("promote"))

            out = config.outputs_dir / manifest["run_id"] / "risks_watchouts.txt"
            self.assertTrue(out.exists())
            text = out.read_text(encoding="utf-8")
            self.assertIn("IT SIMPLIFICATION RISKS & WATCHOUTS", text)
            self.assertIn("END OF REPORT", text)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main(verbosity=2)
